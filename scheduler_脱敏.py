fromfrom __future__ import annotations

import argparse
import math
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from statistics import median
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

SHEET_NAME = "挑战明细"   #保底明细  #冲刺明细  #挑战明细


SHIFT_NAMES = ("白班", "夜班")
FACTORY_SOURCE_MAP = {
    "XX工厂1": "XX1",
    "XX工厂2": "XX2",
    "XX工厂3": "XX3",
    "XX工厂4": "XX4",
}


# 断点SKU集合
BREAKPOINT_SKUS = set()
# 断点SKU映射：{SKU: 工厂}
BREAKPOINT_SKU_FACTORY = {}

@dataclass
class PlanItem:
    month: int
    factory: str
    sku: str
    product_name: str
    category: str
    subcategory: str
    series: str
    spec_piece: int
    planned_qty: int


@dataclass
class CapabilityRow:
    workshop: str
    line_code: str
    line_name: str
    source: str
    category: str
    subcategory: str
    piece_count: int
    speed: float
    target_capacity: int
    shift_minutes: float

def load_breakpoint_skus(path: Path) -> dict:
    """加载断点SKU列表，返回 {SKU: 工厂} 映射"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    breakpoint_map = {}
    
    # 获取表头，找到"物料编码"和"工厂"列
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    sku_col = None
    factory_col = None
    
    for i, header in enumerate(headers):
        if header and "物料编码" in str(header):
            sku_col = i
        elif header and "工厂" in str(header):
            factory_col = i
    
    # 如果找不到列，默认使用第1列和第2列
    if sku_col is None:
        sku_col = 0
    if factory_col is None:
        factory_col = 1
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[sku_col]:
            sku = str(row[sku_col]).strip()
            factory = str(row[factory_col]).strip() if row[factory_col] else ""
            if factory:
                breakpoint_map[sku] = factory
    
    wb.close()
    return breakpoint_map


def infer_target_subcategory(item: PlanItem) -> Optional[str]:
    text = " ".join([item.product_name, item.series, item.category, item.subcategory])
    if "洗浴柔巾" in text or "一次性浴巾" in text:
        return "一次性浴巾"
    return None


def load_transfer_skus(path: Path, target_factory: str) -> dict:
    """加载需要转移工厂的小包10抽SKU列表，返回 {SKU: 目标工厂}"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    transfer_map = {}
    
    # 获取表头
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    sku_col = None
    
    for i, header in enumerate(headers):
        if header and "物料编码" in str(header):
            sku_col = i
            break
    
    if sku_col is None:
        sku_col = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[sku_col]:
            sku = str(row[sku_col]).strip()
            transfer_map[sku] = target_factory
    
    wb.close()
    return transfer_map


def is_strictly_compatible(item: PlanItem, row: CapabilityRow) -> bool:

    global BREAKPOINT_SKU_FACTORY
    
    # ========== 断点SKU严格规则 ==========
    if item.sku in BREAKPOINT_SKU_FACTORY:
        # 必须是断点产线
        if "点断" not in row.line_name and "断点" not in row.line_name:
            return False
        
        # 检查工厂是否匹配
        required_factory = BREAKPOINT_SKU_FACTORY[item.sku]
        if required_factory == "XX工厂4" and row.source != "XX4":
            return False
        if required_factory == "XX工厂1" and row.source != "XX1":
            return False
        
        # 匹配成功，直接返回True
        return True
    
    # 非断点SKU：断点产线不能生产
    if "点断" in row.line_name or "断点" in row.line_name:
        return False
    
    
    # 湿巾/湿厕纸：可以放宽大类限制，但不能跨品类
    wipes_categories = ["湿巾", "湿厕纸"]
    is_wipes_item = item.category in wipes_categories or item.subcategory in wipes_categories
    is_wipes_capability = row.category in wipes_categories or row.subcategory in wipes_categories
    
    # 某种功能湿巾限制：产线含"某种功能湿巾"，物料必须含"某种功能"
    if "某种功能湿巾" in row.line_name:
        if "某种功能" not in item.product_name and "某种功能" not in item.sku and "某种功能" not in item.subcategory:
            return False
    # 2片限制
    if "高速2片" in row.line_name:
        if "2片" not in item.product_name and "2片" not in item.sku and "2片" not in item.subcategory:
            return False
    # 底部抽-悬挂 限制
    if "底部抽" in row.line_name:
        if "悬挂" not in item.product_name and "悬挂" not in item.sku and "悬挂" not in item.subcategory:
            return False
    
    # 100/120抽乳霜纸产线：只能生产含"100抽"或"120抽"的物料
    if "100/120抽乳霜纸" in row.line_name:
        if "100抽" not in item.product_name and "120抽" not in item.product_name:
            return False

    # 单片限制：产线含"单" ↔ 物料含"单片/1片"
    single_keywords = ["单片", "1片"]
    is_single_item = any(kw in item.product_name or kw in item.sku for kw in single_keywords)
    has_single_in_line = "单" in row.line_name or "单" in row.workshop
    
    # 产线有"单"的只能生产单片，产线无"单"的不能生产单片
    if has_single_in_line != is_single_item:
        return False
    
    item_is_panty = item.category == "一次性内裤" or item.subcategory == "一次性内裤"
    if item_is_panty:
        return row.subcategory == "一次性内裤"

    target_subcategory = infer_target_subcategory(item)
    if target_subcategory == "一次性浴巾":
        return row.subcategory == "一次性浴巾"
    
    # 湿巾类：可以放宽大类限制，但必须有对应的品类关系
    if is_wipes_item and is_wipes_capability:
        # 检查具体品类：湿巾产线只生产湿巾，湿厕纸产线只生产湿厕纸
        item_wipes_type = None
        if item.category == "湿巾" or item.subcategory == "湿巾":
            item_wipes_type = "湿巾"
        elif item.category == "湿厕纸" or item.subcategory == "湿厕纸":
            item_wipes_type = "湿厕纸"
        
        row_wipes_type = None
        if row.category == "湿巾" or row.subcategory == "湿巾":
            row_wipes_type = "湿巾"
        elif row.category == "湿厕纸" or row.subcategory == "湿厕纸":
            row_wipes_type = "湿厕纸"
        
        # 品类必须一致
        if item_wipes_type != row_wipes_type:
            return False
        
        return True
    
    # 其他品类：大类必须严格匹配
    if row.category != item.category:
        return False
    
    return True

def is_self_produced(item: PlanItem) -> bool:
    """判断是否为自产品类（非外代工）"""
    # 外代工的SKU列表
    OUTSOURCED_SKUS = {
        "1003051",  # 某款一次性真空洗浴柔巾
        "1002681",  # 某种功能湿巾单片*3包*28提
        "1020682",  # 某种功能湿巾单片*7包
        "1000683",  # 某种功能湿巾单片*192包
        "1002721",  # 某种功能湿巾紫色款单片*3包*40提
        "1020723",  # 某种功能湿巾紫色款单片*7包
        "1000724",  # 某种功能湿巾紫色款单片*192包
        "1020720",  # 某种功能湿巾粉色款单片*3包*40提
        "1002022",  # 某种功能湿巾粉色款单片*7包
        "1000729",  # 某种功能湿巾粉色款单片*192包
        "1002730",  # 某种功能湿巾单片*3包*40提
    }
    
    # 或者根据产品名称判断
    if "某种功能湿巾" in item.product_name and "单片" in item.product_name:
        return False  # 可外代工
        
    if "某款一次性真空洗浴柔巾" in item.product_name:
        return False  # 可外代工
    
    # 默认返回True（自产）
    return True

@dataclass
class LineProfile:
    line_name: str
    workshop: str
    source: str
    base_shift_minutes: float
    capability_rows: List[CapabilityRow] = field(default_factory=list)
    piece_counter: Counter = field(default_factory=Counter)
    category_counter: Counter = field(default_factory=Counter)
    subcategory_counter: Counter = field(default_factory=Counter)
    combo_counter: Counter = field(default_factory=Counter)
    min_piece: int = 999999
    max_piece: int = 0

    def candidate_capability(self, item: PlanItem) -> Optional[Tuple[CapabilityRow, Tuple[float, float, int, int, float]]]:
        best: Optional[Tuple[CapabilityRow, Tuple[float, float, int, int, float]]] = None
        expected_source = FACTORY_SOURCE_MAP.get(item.factory, item.factory)
        target_subcategory = infer_target_subcategory(item)

        # if item.spec_piece == 1:  # 只打印一次，避免刷屏
        #     print(f"产线 {self.line_name}: min_piece={self.min_piece}, max_piece={self.max_piece}, 能力数量={len(self.capability_rows)}")
    
        for row in self.capability_rows:
            # 规格范围检查：SKU的片数必须在产线历史生产过的范围内
            if item.spec_piece < self.min_piece or item.spec_piece > self.max_piece:
                continue  # 超出范围，直接跳过
            
            if not is_strictly_compatible(item, row):
                continue

            score = 0.0
            exact_source = row.source == expected_source
            if exact_source:
                score += 60.0
            else:
                continue

            piece_gap = abs(row.piece_count - item.spec_piece)
            ratio_gap = piece_gap / max(item.spec_piece, 1)
            piece_score = 140.0 if piece_gap == 0 else max(0.0, 110.0 - piece_gap * 2.0 - ratio_gap * 60.0)
            score += piece_score

            exact_subcategory = row.subcategory and row.subcategory == item.subcategory
            exact_category = row.category and row.category == item.category
            target_sub_match = target_subcategory and row.subcategory == target_subcategory
            if target_sub_match:
                score += 220.0
            if exact_subcategory:
                score += 90.0
            elif exact_category:
                score += 45.0
            else:
                score -= 20.0

            if exact_category:
                score += self.category_counter[item.category] * 2.0
            if exact_subcategory:
                score += self.subcategory_counter[item.subcategory] * 4.0
                score += self.combo_counter[(item.category, item.subcategory, row.piece_count)] * 12.0
            elif target_sub_match:
                score += self.subcategory_counter[row.subcategory] * 4.0
                score += self.combo_counter[(row.category, row.subcategory, row.piece_count)] * 12.0
            else:
                score += self.combo_counter[(item.category, row.subcategory, row.piece_count)] * 2.0
            score += self.piece_counter[row.piece_count] * 0.8

            # Same source is mandatory, but avoid overly weak fallbacks unless nothing else exists.
            if piece_score <= 0 and not exact_category and not exact_subcategory and not target_sub_match:
                continue

            tie = (score, row.speed, row.target_capacity, -piece_gap, -ratio_gap)
            if best is None or tie > best[1]:
                best = (row, tie)
        return best
    
@dataclass
class ShiftSlot:
    shift_id: str
    line_name: str
    workshop: str
    shift_date: date
    shift_name: str
    month: int
    is_open_day: bool
    remaining_minutes: float
    fragments: List[dict] = field(default_factory=list)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_int(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0
    return int(round(float(text)))


def as_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(normalize_text(value).replace(",", ""))


def parse_month(value: object) -> int:
    text = normalize_text(value)
    if text.endswith("月"):
        text = text[:-1]
    return int(text)


def choose_sheet_with_headers(workbook, required_headers: Iterable[str]):
    required = set(required_headers)
    best_sheet = None
    best_score = -1
    for ws in workbook.worksheets:
        try:
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        present = {header for header in headers if header is not None}
        score = len(required & present)
        if score > best_score:
            best_score = score
            best_sheet = ws
    if best_sheet is None:
        raise ValueError(f"未找到包含表头 {sorted(required)} 的工作表")
    return best_sheet


def build_capability_from_row(row: tuple, idx: Dict[str, int]) -> CapabilityRow:
    capability = CapabilityRow(
        workshop=normalize_text(row[idx["车间"]]),
        line_code=normalize_text(row[idx["部门编码"]]),
        line_name=normalize_text(row[idx["部门名称"]]),
        source=normalize_text(row[idx["数据来源"]]),
        category=normalize_text(row[idx["大类"]]),
        subcategory=normalize_text(row[idx["二级分类（财务）"]]),
        piece_count=as_int(row[idx["片数"]]),
        speed=as_float(row[idx["机器速度.1"]]),
        target_capacity=as_int(row[idx["目标产能"]]) if "目标产能" in idx else 0,
        shift_minutes=as_float(row[idx["单班有效时长"]]),
    )
    if capability.target_capacity <= 0:
        capability.target_capacity = int(round(capability.speed * capability.shift_minutes))
    return capability


def load_plan_items(path: Path) -> List[PlanItem]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = choose_sheet_with_headers(wb, ["月份", "处理SKU", "规划包数", "规格"])
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {header: i for i, header in enumerate(headers)}
    grouped: Dict[Tuple[int, str, str, str, str, str, int], int] = defaultdict(int)
    names: Dict[Tuple[int, str, str, str, str, str, int], str] = {}
    for row in rows:
        month = parse_month(row[idx["月份"]])
        factory = normalize_text(row[idx["供应工厂"]])
        sku = normalize_text(row[idx["处理SKU"]])
        product_name = normalize_text(row[idx["产品名称"]])
        category = normalize_text(row[idx["大类"]])
        subcategory = normalize_text(row[idx["二级分类"]])
        series = normalize_text(row[idx["系列"]]) if "系列" in idx else ""
        spec_piece = as_int(row[idx["规格"]])
        planned_qty = as_int(row[idx["规划包数"]])
        if not sku or planned_qty <= 0 or spec_piece <= 0:
            continue
        key = (month, factory, sku, category, subcategory, series, product_name, spec_piece)
        grouped[key] += planned_qty
        names[key] = product_name
    wb.close()

    items = [
        PlanItem(
            month=key[0],
            factory=key[1],
            sku=key[2],
            category=key[3],
            subcategory=key[4],
            series=key[5],
            product_name=key[6],
            spec_piece=key[7],
            planned_qty=qty,
        )
        for key, qty in grouped.items()
    ]
    return items


def load_capabilities(path: Path) -> Tuple[List[CapabilityRow], Dict[str, LineProfile]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = choose_sheet_with_headers(wb, ["车间", "部门名称", "机器速度.1", "片数", "数据来源"])
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {header: i for i, header in enumerate(headers)}
    capability_rows: List[CapabilityRow] = []
    profiles: Dict[str, LineProfile] = {}

    for row in rows:
        capability = build_capability_from_row(row, idx)
        line_name = capability.line_name
        if not line_name:
            continue
        if capability.piece_count <= 0 or capability.speed <= 0 or capability.shift_minutes <= 0:
            continue
        capability_rows.append(capability)
        profile = profiles.get(line_name)
        if profile is None:
            profile = LineProfile(
                line_name=line_name,
                workshop=capability.workshop,
                source=capability.source,
                base_shift_minutes=capability.shift_minutes,
            )
            profiles[line_name] = profile
        profile.capability_rows.append(capability)
        profile.piece_counter[capability.piece_count] += 1
        profile.category_counter[capability.category] += 1
        # 添加这两行
        profile.min_piece = min(profile.min_piece, capability.piece_count)
        profile.max_piece = max(profile.max_piece, capability.piece_count)
        profile.subcategory_counter[capability.subcategory] += 1
        profile.combo_counter[(capability.category, capability.subcategory, capability.piece_count)] += 1

    wb.close()

    for profile in profiles.values():
        minute_values = [row.shift_minutes for row in profile.capability_rows if row.shift_minutes > 0]
        profile.base_shift_minutes = float(median(minute_values)) if minute_values else 0.0

    return capability_rows, profiles


def register_capability(capability: CapabilityRow, profiles: Dict[str, LineProfile], capability_rows: Optional[List[CapabilityRow]] = None) -> None:
    if not capability.line_name or capability.piece_count <= 0 or capability.speed <= 0 or capability.shift_minutes <= 0:
        return
    if capability_rows is not None:
        capability_rows.append(capability)
    profile = profiles.get(capability.line_name)
    if profile is None:
        profile = LineProfile(
            line_name=capability.line_name,
            workshop=capability.workshop,
            source=capability.source,
            base_shift_minutes=capability.shift_minutes,
        )
        profiles[capability.line_name] = profile
    profile.capability_rows.append(capability)
    profile.piece_counter[capability.piece_count] += 1
    profile.min_piece = min(profile.min_piece, capability.piece_count)
    profile.max_piece = max(profile.max_piece, capability.piece_count)
    profile.category_counter[capability.category] += 1
    profile.subcategory_counter[capability.subcategory] += 1
    profile.combo_counter[(capability.category, capability.subcategory, capability.piece_count)] += 1
    minute_values = [row.shift_minutes for row in profile.capability_rows if row.shift_minutes > 0]
    profile.base_shift_minutes = float(median(minute_values)) if minute_values else profile.base_shift_minutes


def load_matching_overrides(path: Path) -> Tuple[Dict[Tuple[str, str], List[CapabilityRow]], List[CapabilityRow]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = choose_sheet_with_headers(
        wb,
        ["供应工厂", "处理SKU", "车间", "部门编码", "部门名称", "机器速度.1", "片数", "数据来源", "大类", "二级分类（财务）", "单班有效时长"],
    )
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {header: i for i, header in enumerate(headers)}
    overrides: Dict[Tuple[str, str], List[CapabilityRow]] = defaultdict(list)
    capabilities: List[CapabilityRow] = []
    for row in rows:
        factory = normalize_text(row[idx["供应工厂"]])
        sku = normalize_text(row[idx["处理SKU"]])
        if not factory or not sku:
            continue
        capability = build_capability_from_row(row, idx)
        if not capability.line_name:
            continue
        overrides[(factory, sku)].append(capability)
        capabilities.append(capability)
    wb.close()
    return overrides, capabilities


def load_calendar(path: Path) -> List[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = choose_sheet_with_headers(wb, ["日期", "月", "是否可开工"])
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {header: i for i, header in enumerate(headers)}
    calendar_days: List[dict] = []
    for row in rows:
        raw_date = row[idx["日期"]]
        if isinstance(raw_date, datetime):
            shift_date = raw_date.date()
        elif isinstance(raw_date, date):
            shift_date = raw_date
        else:
            continue
        calendar_days.append(
            {
                "date": shift_date,
                "month": as_int(row[idx["月"]]),
                "is_open_day": normalize_text(row[idx["是否可开工"]]) == "是",
            }
        )
    wb.close()
    return calendar_days


def build_shift_slots(calendar_days: List[dict], profiles: Dict[str, LineProfile]) -> Dict[str, ShiftSlot]:
    slots: Dict[str, ShiftSlot] = {}
    for day in calendar_days:
        for profile in profiles.values():
            for shift_name in SHIFT_NAMES:
                shift_id = f"{day['date'].isoformat()}|{shift_name}|{profile.line_name}"
                slots[shift_id] = ShiftSlot(
                    shift_id=shift_id,
                    line_name=profile.line_name,
                    workshop=profile.workshop,
                    shift_date=day["date"],
                    shift_name=shift_name,
                    month=day["month"],
                    is_open_day=day["is_open_day"],
                    remaining_minutes=profile.base_shift_minutes,
                )
    return slots


def month_sort_key(month: int) -> int:
    return month


def line_shift_phase(slot: ShiftSlot, target_month: int) -> Optional[int]:
    # 当月开工
    if slot.month == target_month and slot.is_open_day:
        return 0
    # 当月休息日
    if slot.month == target_month and not slot.is_open_day:
        return 1
    # 只允许借前一个月的班次（target_month - 1）
    if slot.month == target_month - 1 and slot.is_open_day:
        return 2
    if slot.month == target_month - 1 and not slot.is_open_day:
        return 3
    # 其他月份不允许借用
    return None




def shift_sort_key(slot: ShiftSlot, target_month: int) -> Tuple[int, date, int]:
    phase = line_shift_phase(slot, target_month)
    if phase is None:
        return (99, slot.shift_date, 0)
    # 优先级：当月(0,1) > 前月(2,3)
    if phase in (0, 1):
        return (phase, slot.shift_date, 0 if slot.shift_name == "白班" else 1)
    else:  # 前月
        # 前月的班次按日期正序，越早的越优先用
        return (phase, slot.shift_date, 0 if slot.shift_name == "白班" else 1)

def enumerate_candidate_lines(
    item: PlanItem,
    profiles: Dict[str, LineProfile],
    override_map: Optional[Dict[Tuple[str, str], List[CapabilityRow]]] = None,
) -> List[Tuple[LineProfile, CapabilityRow, Tuple[float, float, int, int, float]]]:
    
    # print(f"白名单keys: {list((override_map or {}).keys())}")
        # 调试文件输出
    debug_file = Path(r"/home/user\Desktop\2026全年规划排产\debug_白名单匹配.txt")
    with open(debug_file, 'a', encoding='utf-8') as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"处理: {item.factory} - {item.sku} - {item.product_name}\n")
        f.write(f"白名单keys: {list((override_map or {}).keys())}\n")
        
        override_rows = (override_map or {}).get((item.factory, item.sku), [])
        f.write(f"找到白名单: {len(override_rows)}条\n")
        
        if override_rows:
            for idx, cap in enumerate(override_rows):
                f.write(f"  白名单{idx+1}: 产线={cap.line_name}, 车间={cap.workshop}\n")
                profile = profiles.get(cap.line_name)
                if profile is None:
                    f.write(f"    ❌ 产线不在profiles中！\n")
                    f.write(f"    profiles中的产线: {list(profiles.keys())}\n")
                else:
                    f.write(f"    ✅ 产线存在\n")
    
    candidates: List[Tuple[LineProfile, CapabilityRow, Tuple[float, float, int, int, float]]] = []
    override_rows = (override_map or {}).get((item.factory, item.sku), [])
    
    # if override_rows:  
    #     for capability in override_rows:
    #         profile = profiles.get(capability.line_name)
    #         if profile is None:
    #             continue
    #         candidates.append((profile, capability, (10_000.0, capability.speed, capability.target_capacity, 0, 0.0)))
    #     candidates.sort(key=lambda entry: entry[2], reverse=True)
    #     return candidates

    if override_rows:
        for capability in override_rows:
            profile = profiles.get(capability.line_name)
            if profile is None:
                # 动态创建profile
                profile = LineProfile(
                    line_name=capability.line_name,
                    workshop=capability.workshop,
                    source=capability.source,
                    base_shift_minutes=capability.shift_minutes,
                )
                profile.capability_rows.append(capability)
                profiles[capability.line_name] = profile
            candidates.append((profile, capability, (10_000.0, capability.speed, capability.target_capacity, 0, 0.0)))
        candidates.sort(key=lambda entry: entry[2], reverse=True)
        return candidates

    for profile in profiles.values():
        match = profile.candidate_capability(item)
        if match is None:
            continue
        capability, score = match
        candidates.append((profile, capability, score))
    candidates.sort(key=lambda entry: entry[2], reverse=True)
    return candidates

    # ==================================================================
    
    candidates.sort(key=lambda entry: entry[2], reverse=True)
    return candidates


def next_shift_for_line(
    line_name: str,
    month: int,
    slots: Dict[str, ShiftSlot],
    sorted_shift_ids_by_line: Dict[str, List[str]],
) -> Optional[ShiftSlot]:
    for shift_id in sorted_shift_ids_by_line[line_name]:
        slot = slots[shift_id]
        if line_shift_phase(slot, month) is None:
            continue
        if slot.remaining_minutes > 0:
            return slot
    return None
    # is_panty_line = line_name == "JQW1-4008-一次性内裤"
    
    # if is_panty_line:
    #     print(f"     查找 {line_name} 在月份 {month} 的班次")
    
    # for shift_id in sorted_shift_ids_by_line[line_name]:
    #     slot = slots[shift_id]
    #     phase = line_shift_phase(slot, month)
    #     if phase is None:
    #         continue
    #     if slot.remaining_minutes > 0:
    #         if is_panty_line:
    #             print(f"       找到班次: {slot.shift_date} {slot.shift_name}, 剩余: {slot.remaining_minutes}")
    #         return slot
    
    # if is_panty_line:
    #     print(f"       没有找到可用班次！")
    # return None

def make_fragment_record(
    month: int,
    item: PlanItem,
    capability: CapabilityRow,
    slot: ShiftSlot,
    qty: int,
    production_minutes: float,
    changeover_minutes: int,
    phase: int,
) -> dict:
    return {
        "月份": month,
        "计划月份": f"{month}月",
        "日期": slot.shift_date,
        "班次": slot.shift_name,
        "班次类型": "正常开工" if slot.is_open_day else "休息日补排",
        "排产层级": ["当月开工", "当月休息日", "前月开工", "前月休息日"][phase],
        "供应工厂": item.factory,
        "车间": capability.workshop,
        "产线": capability.line_name,
        "处理SKU": item.sku,
        "产品名称": item.product_name,
        "大类": item.category,
        "二级分类": item.subcategory,
        "规格": item.spec_piece,
        "机器速度.1": capability.speed,
        "切产损耗分钟": changeover_minutes,
        "生产占用分钟": round(production_minutes, 2),
        "总占用分钟": round(production_minutes + changeover_minutes, 2),
        "本段产量": qty,
        "班次剩余分钟": round(slot.remaining_minutes, 2),
    }


def schedule_month(
    month: int,
    month_items: List[PlanItem],
    profiles: Dict[str, LineProfile],
    slots: Dict[str, ShiftSlot],
    line_slot_index: Dict[str, List[str]],
    changeover_minutes: int,
    override_map: Optional[Dict[Tuple[str, str], List[CapabilityRow]]] = None,
) -> Tuple[List[dict], List[dict], List[dict]]:
    fragments: List[dict] = []
    summary_rows: List[dict] = []
    unmatched_rows: List[dict] = []
    line_used_minutes: Dict[str, float] = defaultdict(float)
    line_fragments: Dict[str, int] = defaultdict(int)
    sorted_shift_ids_by_line: Dict[str, List[str]] = {}
    for line_name, shift_ids in line_slot_index.items():
        sorted_shift_ids_by_line[line_name] = sorted(
            shift_ids,
            key=lambda shift_id: shift_sort_key(slots[shift_id], month),
        )


    sorted_items = sorted( 
        month_items,
        key=lambda item: (
            # 优先级0：是否可外代工（0=自产优先，1=外代工排后）
            0 if is_self_produced(item) else 1,
            # 优先级1：规划数量大的优先
            -item.planned_qty,
            # 优先级2：工厂
            item.factory,
            # 优先级3：品类
            item.category,
            # 优先级4：子分类
            item.subcategory,
        ),
    )
        
    for item in sorted_items:
        remaining_qty = item.planned_qty

        candidates = enumerate_candidate_lines(item, profiles, override_map)
        if not candidates:
            unmatched_rows.append(
                {
                    "月份": month,
                    "供应工厂": item.factory,
                    "处理SKU": item.sku,
                    "产品名称": item.product_name,
                    "大类": item.category,
                    "二级分类": item.subcategory,
                    "规格": item.spec_piece,
                    "规划包数": item.planned_qty,
                    "原因": "没有找到匹配片数和工厂来源的产线",
                }
            )
            summary_rows.append(
                {
                    "月份": month,
                    "供应工厂": item.factory,
                    "处理SKU": item.sku,
                    "产品名称": item.product_name,
                    "大类": item.category,
                    "二级分类": item.subcategory,
                    "规格": item.spec_piece,
                    "规划包数": item.planned_qty,
                    "实际排产合计": 0,
                    "差异": item.planned_qty,
                    "状态": "未匹配",
                }
            )
            continue

        actual_qty = 0
        candidate_lookup = {
            profile.line_name: (profile, capability, score)
            for profile, capability, score in candidates
        }
        while remaining_qty > 0:
            options = []
            for profile, capability, score in candidates:
                slot = next_shift_for_line(profile.line_name, month, slots, sorted_shift_ids_by_line)
                if slot is None:
                    continue
                phase = line_shift_phase(slot, month)
                if phase is None:
                    continue
                unused_bonus = 100000 if line_fragments[profile.line_name] == 0 else 0
                load_penalty = line_used_minutes[profile.line_name]
                #计算优先级元组，班次剩余时间少的，靠后排
                options.append(
                    (
                        phase,
                        -unused_bonus,
                        load_penalty,   #符合条件的产线轮流使用，平衡产线负荷状况
                        -score[0],
                        -capability.speed,
                        slot.shift_date,
                        0 if slot.shift_name == "白班" else 1,
                        profile.line_name,
                    )
                )

            if not options:
                break

            options.sort() # ← 贪心选择：选当前最优
            chosen_line_name = options[0][-1]
            profile, capability, _score = candidate_lookup[chosen_line_name]
            slot = next_shift_for_line(chosen_line_name, month, slots, sorted_shift_ids_by_line)
            if slot is None:
                break
            phase = line_shift_phase(slot, month)
            if phase is None:
                break
            if slot.remaining_minutes <= 0:
                continue

            needs_changeover = len(slot.fragments) > 0
            fixed_loss = changeover_minutes if needs_changeover else 0
            usable_minutes = slot.remaining_minutes - fixed_loss
            if usable_minutes <= 0:
                slot.remaining_minutes = 0.0
                continue

            max_qty = int(math.floor(capability.speed * usable_minutes))
            if max_qty <= 0:
                slot.remaining_minutes = 0.0
                continue

            qty = min(remaining_qty, max_qty)
            production_minutes = qty / capability.speed
            consumed_minutes = production_minutes + fixed_loss

            if consumed_minutes > slot.remaining_minutes + 1e-9:
                qty = int(math.floor(capability.speed * max(slot.remaining_minutes - fixed_loss, 0)))
                if qty <= 0:
                    slot.remaining_minutes = 0.0
                    continue
                production_minutes = qty / capability.speed
                consumed_minutes = production_minutes + fixed_loss

            slot.remaining_minutes = max(slot.remaining_minutes - consumed_minutes, 0.0)
            actual_qty += qty
            remaining_qty -= qty
            line_used_minutes[chosen_line_name] += consumed_minutes
            line_fragments[chosen_line_name] += 1

            record = make_fragment_record(
                month=month,
                item=item,
                capability=capability,
                slot=slot,
                qty=qty,
                production_minutes=production_minutes,
                changeover_minutes=fixed_loss,
                phase=phase,
            )
            slot.fragments.append(record)
            fragments.append(record)

        status = "完成"
        if actual_qty == 0:
            status = "未排"
        elif remaining_qty > 0:
            status = "部分完成"

        summary_rows.append(
            {
                "月份": month,
                "供应工厂": item.factory,
                "处理SKU": item.sku,
                "产品名称": item.product_name,
                "大类": item.category,
                "二级分类": item.subcategory,
                "规格": item.spec_piece,
                "规划包数": item.planned_qty,
                "实际排产合计": actual_qty,
                "差异": item.planned_qty - actual_qty,
                "状态": status,
            }
        )

        if remaining_qty > 0:
            unmatched_rows.append(
                {
                    "月份": month,
                    "供应工厂": item.factory,
                    "处理SKU": item.sku,
                    "产品名称": item.product_name,
                    "大类": item.category,
                    "二级分类": item.subcategory,
                    "规格": item.spec_piece,
                    "规划包数": item.planned_qty,
                    "未完成数量": remaining_qty,
                    "原因": "当月开工 + 当月休息日 + 前月借班次后仍不足",
                }
            )

    return fragments, summary_rows, unmatched_rows


def auto_fit_sheet(ws) -> None:
    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), width)
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 28)


def build_month_layout_rows(month: int, fragments: List[dict], summary_rows: List[dict], calendar_days: List[dict]) -> Tuple[List[dict], Dict[str, bool]]:
    month_days = [day for day in calendar_days if day["month"] == month]
    shift_headers: List[str] = []
    shift_open_map: Dict[str, bool] = {}
    for day in month_days:
        for shift_name in SHIFT_NAMES:
            header = f"{day['date'].day}日{shift_name}"
            shift_headers.append(header)
            shift_open_map[header] = day["is_open_day"]

    summary_lookup = {row["处理SKU"]: row for row in summary_rows}
    layout_rows: List[dict] = []

    ordered_fragments = sorted(
        fragments,
        key=lambda row: (
            row["车间"],
            row["产线"],
            row["日期"],
            0 if row["班次"] == "白班" else 1,
            row["处理SKU"],
        ),
    )

    for record in ordered_fragments:
        row = {
            "处理SKU": record["处理SKU"],
            "产品名称": record["产品名称"],
            "供应工厂": record["供应工厂"],
            "车间": record["车间"],
            "产线": record["产线"],
            "规格": record["规格"],
            "月度规划": summary_lookup.get(record["处理SKU"], {}).get("规划包数", record["本段产量"]),
        }
        for header in shift_headers:
            row[header] = None
        target_header = f"{record['日期'].day}日{record['班次']}"
        if target_header in row:
            row[target_header] = record["本段产量"]
        row["实际排产合计"] = record["本段产量"]
        row["排产层级"] = record["排产层级"]
        row["切产损耗分钟"] = record["切产损耗分钟"]
        row["生产占用分钟"] = record["生产占用分钟"]
        layout_rows.append(row)

    return layout_rows, shift_open_map


def append_sheet(wb: Workbook, title: str, rows: List[dict]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header)
            if isinstance(value, date):
                values.append(value.strftime("%Y-%m-%d"))
            else:
                values.append(value)
        ws.append(values)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_sheet(ws)


def append_month_layout_sheet(
    wb: Workbook,
    title: str,
    rows: List[dict],
    shift_open_map: Dict[str, bool],
) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(header) for header in headers])

    gray_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for col_idx, header in enumerate(headers, start=1):
        if header in shift_open_map and not shift_open_map[header]:
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for item in cell:
                    item.fill = gray_fill

    ws.freeze_panes = "H2"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_sheet(ws)


def export_workbook(
    output_path: Path,
    month_results: Dict[int, dict],
    metadata_rows: List[dict],
    calendar_days: List[dict],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    append_sheet(wb, "参数说明", metadata_rows)

    for month in sorted(month_results):
        result = month_results[month]
        layout_rows, shift_open_map = build_month_layout_rows(month, result["fragments"], result["summary"], calendar_days)
        append_month_layout_sheet(wb, f"{month:02d}月排产表", layout_rows, shift_open_map)
        append_sheet(wb, f"{month:02d}月明细", result["fragments"])
        append_sheet(wb, f"{month:02d}月汇总", result["summary"])

    wb.save(output_path)


def build_line_slot_index(slots: Dict[str, ShiftSlot]) -> Dict[str, List[str]]:
    index: Dict[str, List[str]] = defaultdict(list)
    for shift_id, slot in slots.items():
        index[slot.line_name].append(shift_id)
    return index



def export_capacity_gap_report(output_path: Path, month_results: Dict[int, dict]) -> None:
    """导出产能缺口分析报表（独立文件）"""
    
    # ========== 全年生产汇总（无论是否有缺口都要生成） ==========
    yearly_summary = defaultdict(lambda: {"planned": 0, "actual": 0})
    
    for month, result in month_results.items():
        for summary in result["summary"]:
            key = (summary["月份"], summary["供应工厂"], summary["大类"])
            yearly_summary[key]["planned"] += summary["规划包数"]
            yearly_summary[key]["actual"] += summary["实际排产合计"]
    
    # ========== 收集缺口数据 ==========
    all_gaps = []
    
    for month, result in month_results.items():
        for summary in result["summary"]:
            if summary["差异"] > 0:
                # 判断缺口原因
                if summary["状态"] == "未匹配":
                    reason = "没有找到匹配的产线"
                elif summary["状态"] == "未排":
                    reason = "有匹配产线但无可用班次"
                elif summary["状态"] == "部分完成":
                    reason = "产能不足，仅部分完成"
                else:
                    reason = "其他原因"
                
                all_gaps.append({
                    "月份": summary["月份"],
                    "供应工厂": summary["供应工厂"],
                    "处理SKU": summary["处理SKU"],
                    "产品名称": summary["产品名称"],
                    "大类": summary["大类"],
                    "二级分类": summary["二级分类"],
                    "规格": summary["规格"],
                    "规划数量": summary["规划包数"],
                    "已排数量": summary["实际排产合计"],
                    "缺口数量": summary["差异"],
                    "完成率": f"{summary['实际排产合计']/summary['规划包数']*100:.1f}%" if summary['规划包数'] > 0 else "0%",
                    "状态": summary["状态"],
                    "缺口原因": reason,
                })
    
    # 创建新工作簿
    wb = Workbook()
    
    # ========== Sheet1: 全年生产汇总（放在第一个） ==========
    ws_yearly = wb.active
    ws_yearly.title = "全年生产汇总"
    ws_yearly.append(["月份", "供应工厂", "大类", "规划包数", "实际生产包数", "完成率"])
    
    for (month, factory, category), data in sorted(yearly_summary.items()):
        planned = data["planned"]
        actual = data["actual"]
        rate = f"{actual/planned*100:.1f}%" if planned > 0 else "0%"
        ws_yearly.append([month, factory, category, planned, actual, rate])
    
    # 如果没有缺口，只输出全年生产汇总
    if not all_gaps:
        print("✅ 无产能缺口，所有规划均可完成")
        # 自动调整列宽
        for col_idx, col in enumerate(ws_yearly.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_yearly.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)
        wb.save(output_path)
        print(f"\n📊 全年生产汇总已导出: {output_path}")
        return
    
    # ========== Sheet2: 缺口明细 ==========
    ws_detail = wb.create_sheet("缺口明细")
    headers = list(all_gaps[0].keys())
    ws_detail.append(headers)
    for gap in all_gaps:
        ws_detail.append([gap[h] for h in headers])
    
    # ========== Sheet3: 按品类汇总 ==========
    ws_category = wb.create_sheet("按品类汇总")
    category_gaps = defaultdict(int)
    category_planned = defaultdict(int)
    category_actual = defaultdict(int)
    
    for gap in all_gaps:
        cat = gap["大类"]
        category_gaps[cat] += gap["缺口数量"]
        category_planned[cat] += gap["规划数量"]
        category_actual[cat] += gap["已排数量"]
    
    ws_category.append(["大类", "总规划数量", "总已排数量", "总缺口数量", "完成率", "建议"])
    for cat in sorted(category_gaps.keys(), key=lambda x: -category_gaps[x]):
        planned = category_planned[cat]
        actual = category_actual[cat]
        gap = category_gaps[cat]
        rate = f"{actual/planned*100:.1f}%" if planned > 0 else "0%"
        
        if gap > 100000:
            suggestion = "🚨 严重不足，建议立即增加产线"
        elif gap > 50000:
            suggestion = "⚠️ 明显不足，建议评估增加产线"
        elif gap > 10000:
            suggestion = "📌 轻度不足，建议关注"
        else:
            suggestion = "✅ 基本充足"
        
        ws_category.append([cat, planned, actual, gap, rate, suggestion])
    
    # ========== Sheet4: 按SKU汇总（TOP30） ==========
    ws_sku = wb.create_sheet("TOP30缺口SKU")
    ws_sku.append(["排名", "处理SKU", "产品名称", "大类", "规格", "总缺口数量", "建议"])
    
    sku_gaps = defaultdict(lambda: {"planned": 0, "actual": 0})
    for gap in all_gaps:
        sku = gap["处理SKU"]
        sku_gaps[sku]["planned"] += gap["规划数量"]
        sku_gaps[sku]["actual"] += gap["已排数量"]
        sku_gaps[sku]["name"] = gap["产品名称"]
        sku_gaps[sku]["category"] = gap["大类"]
        sku_gaps[sku]["spec"] = gap["规格"]
    
    sku_list = []
    for sku, data in sku_gaps.items():
        sku_list.append({
            "sku": sku,
            "name": data["name"],
            "category": data["category"],
            "spec": data["spec"],
            "gap": data["planned"] - data["actual"]
        })
    sku_list.sort(key=lambda x: -x["gap"])
    
    for idx, sku in enumerate(sku_list[:30], 1):
        if sku["gap"] > 50000:
            suggestion = "建议自建产线或外协"
        elif sku["gap"] > 20000:
            suggestion = "建议增加班次或外协"
        else:
            suggestion = "建议调整排产优先级"
        
        ws_sku.append([idx, sku["sku"], sku["name"], sku["category"], 
                       sku["spec"], sku["gap"], suggestion])
    
    # ========== Sheet5: 按月份汇总 ==========
    ws_month = wb.create_sheet("按月份汇总")
    month_gaps = defaultdict(int)
    for gap in all_gaps:
        month_gaps[gap["月份"]] += gap["缺口数量"]
    
    ws_month.append(["月份", "缺口总数", "建议"])
    for month in sorted(month_gaps.keys()):
        gap = month_gaps[month]
        if gap > 50000:
            suggestion = "本月产能严重不足，建议提前备产"
        elif gap > 20000:
            suggestion = "本月产能不足，建议分析原因"
        else:
            suggestion = "产能基本满足"
        ws_month.append([f"{month}月", gap, suggestion])
    
    # 自动调整列宽
    for ws in [ws_yearly, ws_detail, ws_category, ws_sku, ws_month]:
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)
    
    # 保存文件
    wb.save(output_path)
    print(f"\n📊 产能缺口分析已导出: {output_path}")
    print(f"   - 缺口品类: {len(category_gaps)} 个")
    print(f"   - 缺口SKU: {len(sku_gaps)} 个")
    print(f"   - 总缺口数量: {sum(gap['缺口数量'] for gap in all_gaps):,}")

def export_utilization_report(output_path: Path, month_results: Dict[int, dict], calendar_days: List[dict]) -> None:
    """导出设备开机率报表（多Sheet：产线明细 + 车间汇总 + 工厂汇总 + 月度总览）"""
    
    # 存储所有明细数据
    detail_rows = []
    
    for month in sorted(month_results.keys()):
        result = month_results[month]
        
        # 获取该月日历信息
        month_days = [day for day in calendar_days if day["month"] == month and day["is_open_day"]]
        work_days = len(month_days)
        available_shifts_per_line = work_days * 2
        
        # 统计每条产线的实际开班次数（去重）
        line_unique_shifts = defaultdict(set)
        line_production_minutes = defaultdict(int)
        line_workshop = {}
        line_factory = {}
        
        for fragment in result["fragments"]:
            line_name = fragment["产线"]
            shift_key = f"{fragment['日期']}_{fragment['班次']}"
            line_unique_shifts[line_name].add(shift_key)
            line_production_minutes[line_name] += fragment.get("生产占用分钟", 0)
            
            if line_name not in line_workshop:
                line_workshop[line_name] = fragment.get("车间", "")
                line_factory[line_name] = fragment.get("供应工厂", "")
        
        # 生成明细行
        for line_name in line_unique_shifts.keys():
            actual_shifts = len(line_unique_shifts[line_name])
            utilization = actual_shifts / available_shifts_per_line * 100 if available_shifts_per_line > 0 else 0
            
            if utilization >= 95:
                status = "满负荷"
            elif utilization >= 70:
                status = "正常"
            elif utilization >= 30:
                status = "偏低"
            else:
                status = "闲置"
            
            detail_rows.append({
                "月份": month,
                "工厂": line_factory.get(line_name, ""),
                "车间": line_workshop.get(line_name, ""),
                "产线": line_name,
                "可开工天数": work_days,
                "可开工班次": available_shifts_per_line,
                "实际开班": actual_shifts,
                "开机率(%)": round(utilization, 1),
                "生产分钟": round(line_production_minutes[line_name], 0),
                "状态": status,
            })
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # ========== Sheet1: 按产线明细 ==========
    ws_detail = wb.active
    ws_detail.title = "按产线明细"
    detail_headers = ["月份", "工厂", "车间", "产线", "可开工天数", "可开工班次", "实际开班", "开机率(%)", "生产分钟", "状态"]
    ws_detail.append(detail_headers)
    for row in detail_rows:
        ws_detail.append([row["月份"], row["工厂"], row["车间"], row["产线"], 
                         row["可开工天数"], row["可开工班次"], row["实际开班"],
                         row["开机率(%)"], row["生产分钟"], row["状态"]])
    
    # ========== Sheet2: 按车间汇总 ==========
    ws_workshop = wb.create_sheet("按车间汇总")
    workshop_stats = defaultdict(lambda: {"available": 0, "actual": 0, "minutes": 0, "line_count": 0})
    for row in detail_rows:
        key = (row["月份"], row["工厂"], row["车间"])
        workshop_stats[key]["available"] += row["可开工班次"]
        workshop_stats[key]["actual"] += row["实际开班"]
        workshop_stats[key]["minutes"] += row["生产分钟"]
        workshop_stats[key]["line_count"] += 1
    
    ws_workshop.append(["月份", "工厂", "车间", "产线数量", "可开工班次", "实际开班", "开机率(%)", "生产分钟"])
    for (month, factory, workshop), stats in sorted(workshop_stats.items()):
        utilization = stats["actual"] / stats["available"] * 100 if stats["available"] > 0 else 0
        ws_workshop.append([month, factory, workshop, stats["line_count"], 
                           stats["available"], stats["actual"], round(utilization, 1), round(stats["minutes"], 0)])
    
    # ========== Sheet3: 按工厂汇总 ==========
    ws_factory = wb.create_sheet("按工厂汇总")
    factory_stats = defaultdict(lambda: {"available": 0, "actual": 0, "minutes": 0, "line_count": 0})
    for row in detail_rows:
        key = (row["月份"], row["工厂"])
        factory_stats[key]["available"] += row["可开工班次"]
        factory_stats[key]["actual"] += row["实际开班"]
        factory_stats[key]["minutes"] += row["生产分钟"]
        factory_stats[key]["line_count"] += 1
    
    ws_factory.append(["月份", "工厂", "产线数量", "可开工班次", "实际开班", "开机率(%)", "生产分钟"])
    for (month, factory), stats in sorted(factory_stats.items()):
        utilization = stats["actual"] / stats["available"] * 100 if stats["available"] > 0 else 0
        ws_factory.append([month, factory, stats["line_count"], 
                          stats["available"], stats["actual"], round(utilization, 1), round(stats["minutes"], 0)])
    
    # ========== Sheet4: 月度总览 ==========
    ws_monthly = wb.create_sheet("月度总览")
    monthly_stats = defaultdict(lambda: {"available": 0, "actual": 0, "minutes": 0})
    for row in detail_rows:
        monthly_stats[row["月份"]]["available"] += row["可开工班次"]
        monthly_stats[row["月份"]]["actual"] += row["实际开班"]
        monthly_stats[row["月份"]]["minutes"] += row["生产分钟"]
    
    ws_monthly.append(["月份", "可开工班次", "实际开班", "开机率(%)", "生产分钟"])
    for month in sorted(monthly_stats.keys()):
        stats = monthly_stats[month]
        utilization = stats["actual"] / stats["available"] * 100 if stats["available"] > 0 else 0
        ws_monthly.append([month, stats["available"], stats["actual"], round(utilization, 1), round(stats["minutes"], 0)])
    
    # 自动调整列宽
    for ws in [ws_detail, ws_workshop, ws_factory, ws_monthly]:
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 25)
    
    # 保存
    output_file = output_path.parent / f"{output_path.stem}_开机率分析{output_path.suffix}"
    wb.save(output_file)
    print(f"📊 开机率分析已导出: {output_file}")

def infer_input_paths(base_dir: Path) -> Dict[str, Path]:
    from pathlib import Path

    
    calendar = Path(r"/home/user\Desktop\2026全年规划排产\2026工厂生产日历.xlsx")
    capacity = Path(r"/home/user\Desktop\2026全年规划排产\产线产能_目标产能计算完成.xlsx")
    plan = Path(fr"/home/user\Desktop\2026全年规划排产\目标规划_无外代工_{SHEET_NAME}.xlsx")
    matching_override = Path(r"/home/user\Desktop\2026全年规划排产\匹配补充.xlsx")

    return {
        "calendar": calendar,
        "capacity": capacity,
        "plan": plan,
        "matching_override": matching_override,
    }

def main() -> None:
    #清空调试文件
    global BREAKPOINT_SKU_FACTORY
    
    debug_file = Path(r"/home/user\Desktop\2026全年规划排产\debug_白名单匹配.txt")
    if debug_file.exists():
        debug_file.unlink()

    parser = argparse.ArgumentParser(description="2026全年排产脚本")
    parser.add_argument("--base-dir", default=".", help="源文件所在目录")
    
    output_path = fr"/home/user\Desktop\2026全年规划排产\2026全年排产结果_{SHEET_NAME}.xlsx"
    parser.add_argument("--output", default=output_path, help="输出文件名")
    #parser.add_argument("--output", default=r"/home/user\Desktop\2026全年规划排产\2026全年排产结果_{SHEET_NAME}.xlsx", help="输出文件名")
    parser.add_argument("--changeover-minutes", type=int, default=10, help="切换SKU时扣减的分钟数")
    args = parser.parse_args()
    # gap_output_path = base_dir / f"{args.output.replace('.xlsx', '')}_产能缺口分析_{SHEET_NAME}.xlsx"
    # export_capacity_gap_report(gap_output_path, month_results)

    base_dir = Path(args.base_dir).resolve()
    input_paths = infer_input_paths(base_dir)

    plan_items = load_plan_items(input_paths["plan"])
    _capability_rows, profiles = load_capabilities(input_paths["capacity"])

    # 1.加载断点SKU映射
    breakpoint_path = Path(r"/home/user\Desktop\2026全年规划排产\断点SKU.xlsx")
    if breakpoint_path.exists():
        BREAKPOINT_SKU_FACTORY = load_breakpoint_skus(breakpoint_path)
        print(f"加载断点SKU: {len(BREAKPOINT_SKU_FACTORY)} 个")
        
        # 修改断点SKU的供应工厂
        for item in plan_items:
            if item.sku in BREAKPOINT_SKU_FACTORY:
                new_factory = BREAKPOINT_SKU_FACTORY[item.sku]
                if item.factory != new_factory:
                    print(f"修改SKU工厂: {item.sku} {item.factory} → {new_factory}")
                    item.factory = new_factory

  # 2. 加载小包SKU（也转移到XX4）
    small_pack_path = Path(r"/home/user\Desktop\2026全年规划排产\小包SKU.xlsx")
    if small_pack_path.exists():
        small_pack_map = load_transfer_skus(small_pack_path, "XX工厂4")
        print(f"加载小包SKU: {len(small_pack_map)} 个")
        for item in plan_items:
            if item.sku in small_pack_map:
                item.factory = small_pack_map[item.sku]


    # 加载并注册 override 能力
    override_map, override_capabilities = load_matching_overrides(input_paths["matching_override"])
    
    # 将 override 能力注册到 profiles 中
    for capability in override_capabilities:
        register_capability(capability, profiles)



        # 注册白名单到 profiles
    for capability in override_capabilities:
        register_capability(capability, profiles)
    

    calendar_days = load_calendar(input_paths["calendar"])
    slots = build_shift_slots(calendar_days, profiles)
    line_slot_index = build_line_slot_index(slots)

    # # 调试：打印所有产线名称
    # print("\n=== 所有产线 (profiles) ===")
    # for line_name in profiles.keys():
    #     print(f"  {line_name}")
    
    # 调试：检查白名单产线是否在 profiles 中
    # print("\n=== 检查白名单产线 ===")
    # for capability in override_capabilities:
    #     if capability.line_name in profiles:
    #         print(f"✅ {capability.line_name} 在 profiles 中")
    #     else:
    #         print(f"❌ {capability.line_name} 不在 profiles 中！")
    
    # 创建班次槽位
    slots = build_shift_slots(calendar_days, profiles)
    
    # # 调试：检查班次槽位中的产线
    # print("\n=== 班次槽位中的产线 ===")
    # slot_lines = set()
    # for slot in slots.values():
    #     slot_lines.add(slot.line_name)
    # for line_name in sorted(slot_lines):
    #     print(f"  {line_name}")
    
    # 检查白名单产线是否有班次
    # print("\n=== 白名单产线的班次槽位 ===")
    # for capability in override_capabilities:
    #     has_slot = any(slot.line_name == capability.line_name for slot in slots.values())
    #     if has_slot:
    #         print(f"✅ {capability.line_name} 有班次槽位")
    #         # 统计班次数量
    #         slot_count = sum(1 for slot in slots.values() if slot.line_name == capability.line_name)
    #         print(f"   班次数量: {slot_count}")
    #     else:
    #         print(f"❌ {capability.line_name} 没有班次槽位！")



    month_groups: Dict[int, List[PlanItem]] = defaultdict(list)
    for item in plan_items:
        month_groups[item.month].append(item)

    month_results: Dict[int, dict] = {}
    for month in sorted(month_groups.keys()): #, reverse=True   #改成正序排产
        fragments, summary_rows, unmatched_rows = schedule_month(
            month=month,
            month_items=month_groups[month],
            profiles=profiles,
            slots=slots,
            line_slot_index=line_slot_index,
            changeover_minutes=args.changeover_minutes,
            override_map=override_map, #调用传递
        )
        month_results[month] = {
            "fragments": fragments,
            "summary": summary_rows,
            "exceptions": unmatched_rows,
        }
        completed = sum(1 for row in summary_rows if row["状态"] == "完成")
        print(
            f"{month:02d}月: SKU {len(summary_rows)} 个, 完成 {completed} 个, "
            f"异常 {len(unmatched_rows)} 条, 排产片段 {len(fragments)} 条"
        )

    metadata_rows = [
        {"参数": "切产损耗分钟", "值": args.changeover_minutes, "说明": "同一班次从上一个SKU切到下一个SKU时扣减"},
        {"参数": "排产顺序", "值": "按月度规划包数从大到小", "说明": "月份按12月到1月倒序处理，支持后月向前借班次"},
        {"参数": "班次启用优先级", "值": "当月开工 -> 当月休息日 -> 前月开工 -> 前月休息日", "说明": "A后B"},
        {"参数": "匹配逻辑", "值": "工厂来源优先，二级分类优先，大类兜底，规格相近优先", "说明": "允许多大类多二级分类产线参与"},
        {"参数": "输入目录", "值": str(base_dir), "说明": "自动识别3个源文件"},
    ]
    output_path = base_dir / args.output
    export_workbook(output_path, month_results, metadata_rows, calendar_days)


    # 导出缺口分析报表
    gap_output_path = base_dir / f"{args.output.replace('.xlsx', '')}_缺口分析.xlsx"
    export_capacity_gap_report(gap_output_path, month_results)
    #导出开机率表
    export_utilization_report(output_path, month_results, calendar_days)

    print(f"输出完成: {output_path}")


if __name__ == "__main__":
    main()
